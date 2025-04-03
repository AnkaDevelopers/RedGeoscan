# Importar Modulos
from config import config

# Importaciones 
from openpyxl import Workbook
import pandas as pd
import openpyxl
import os

# Variables globales
path_excel = config.ruta_excel 

#************************************************************************************************************
# Función para crear o abrir el archivo Excel
def inicializar_excel(nombre_archivo):
    
    # Unimos la ruta base 'path_excel' con el nombre del archivo
    path_completo = os.path.join(path_excel, nombre_archivo)
    
    # Validamos si la ruta con el archivo existe
    if not os.path.exists(path_completo):
        
        # Si el archivo no existe, lo creamos con encabezados
        wb = Workbook()
        ws = wb.active
        ws.title = "Rutas Capturadas"
        ws.append(["Fecha", "Nombre del Remitente","Nombre Proyecto", "Ruta Capturada", "ID del Correo", "Estado","Revisado","Estado_chc"])
        wb.save(path_completo)
    
    # En caso de que el archivo ya exsita no pasa nada
    print(f"Archivo Excel inicializado: {nombre_archivo}")
    
    return path_completo


#************************************************************************************************************
# Función para guardar los datos en Excel
def guardar_en_excel(ruta_archivo_excel, fecha, remitente, ruta, mensaje_id, nombre_proyecto, estado):
    
    try:
        
        # Caragmos el archivo excel y guardamos la informacion en las columnas
        wb = openpyxl.load_workbook(ruta_archivo_excel)
        ws = wb.active
        ws.append([fecha, remitente, nombre_proyecto, ruta, mensaje_id, estado, '',estado])
        wb.save(ruta_archivo_excel)
        
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")
        
#************************************************************************************************************
# Función para verificar si un ID ya está en el Excel
def id_ya_registrado(ruta_archivo_excel, mensaje_id):

    try:
        
        wb = openpyxl.load_workbook(ruta_archivo_excel)
        ws = wb.active
        
        for fila in ws.iter_rows(min_row=2, max_col=6, values_only=True):  # Ignorar encabezados
            if fila[4] == mensaje_id:  # Compara con la columna de IDs
                return True
        return False
    except Exception as e:
        print(f"Error al verificar ID en Excel: {e}")
        return False
    
    
#************************************************************************************************************
# Función para verificar si una ruta ya está en el Excel
def ruta_ya_en_cola(ruta_archivo_excel, ruta_proyecto):

    try:        
        wb = openpyxl.load_workbook(ruta_archivo_excel)
        ws = wb.active
        
        for fila in ws.iter_rows(min_row=2, max_col=6, values_only=True):  # Ignorar encabezados
            if fila[3] == ruta_proyecto:  # Compara con la columna de Rutas Capturadas
                return True, fila[0], fila[1]  # Retorna True y la fecha de registro
            
        return False, None, None
    
    except Exception as e:
        print(f"Error al verificar si la ruta ya está en cola: {e}")
        return False, None, None

#************************************************************************************************************
def obtener_primera_fecha(ruta_archivo_excel):
    remitente = []
    
    try:
        # Leer el archivo Excel
        datos_cola = pd.read_excel(ruta_archivo_excel)

        # Validar que el archivo Excel no esté vacío
        if not datos_cola.empty:
            # Convertir la columna 'Fecha' a formato datetime
            datos_cola['Fecha'] = pd.to_datetime(datos_cola['Fecha'], errors='coerce')

            # Filtrar filas con fechas válidas
            datos_validos = datos_cola.dropna(subset=['Fecha'])

            if not datos_validos.empty:
                # Obtener la fecha más antigua
                primera_fecha = datos_validos['Fecha'].min()

                # Obtener la fila correspondiente a la primera fecha
                primera_fila = datos_validos.loc[datos_validos['Fecha'] == primera_fecha].iloc[0]
                primer_nombre = primera_fila['Nombre del Remitente']

                print(f"La primera fecha en el archivo es: {primera_fecha}")
                print(f"El nombre del remitente es: {primer_nombre}")
                
                # Agregar datos al remitente
                remitente.append({"Fecha": primera_fecha, "Nombre del Remitente": primer_nombre})

                print(remitente)
            else:
                print("No se encontraron fechas válidas en el archivo.")
        else:
            print("El archivo Excel está vacío.")
    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")

# Ruta del archivo de prueba
ruta_prueba = r"C:\bot-auto\docs\rutas_redGeoScan.xlsx"
#obtener_primera_fecha(ruta_prueba)