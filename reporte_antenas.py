import os
import openpyxl
from openpyxl.styles import PatternFill

def info_report_antenas(datos_json, ruta_base, carpeta_nombre):
    try:
        # Crear la ruta de la carpeta para guardar el archivo
        ruta_carpeta = os.path.join(ruta_base, carpeta_nombre)
        if not os.path.exists(ruta_carpeta):
            os.makedirs(ruta_carpeta)
            print(f"Carpeta creada: {ruta_carpeta}")
        else:
            print(f"Carpeta ya existente: {ruta_carpeta}")

        # Definir la ruta del archivo Excel
        archivo_excel = os.path.join(ruta_carpeta, f"{carpeta_nombre}.xlsx")

        # Crear un libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Antenas"

        # Escribir encabezados
        encabezados = ["NAME", "Latitud (Decimal)", "Longitud (Decimal)", "Distancia", 
                       "ADMINISTRADOR", "has_rinex", "estado_coordenada", "coordenada"]
        ws.append(encabezados)

        # Crear estilos para las filas
        fill_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde más intenso
        fill_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo más intenso

        # Procesar y escribir los datos
        for item in datos_json:
            # Determinar el color según las condiciones
            has_rinex = item.get("has_rinex", False)
            fill = fill_verde if has_rinex else fill_rojo

            # Formatear la distancia
            distancia_formateada = f"{item.get('Distancia'):.1f} km" if item.get("Distancia") else "N/A"

            # Escribir la fila
            fila = [
                item.get("NAME"),
                item.get("Latitud (Decimal)"),
                item.get("Longitud (Decimal)"),
                distancia_formateada,
                item.get("ADMINISTRADOR"),
                item.get("has_rinex"),
                item.get("estado_coordenada"),
                item.get("coordenada")
            ]
            ws.append(fila)

            # Aplicar el color a toda la fila
            fila_actual = ws.max_row
            for col in range(1, len(fila) + 1):  # Desde la primera hasta la última columna
                ws.cell(row=fila_actual, column=col).fill = fill

        # Ajustar el ancho de las columnas automáticamente
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_length + 2, 15)  # Ajustar al contenido o un mínimo

        # Guardar el archivo Excel
        wb.save(archivo_excel)
        print(f"Archivo Excel creado en: {archivo_excel}")

    except Exception as e:
        print(f"Error al generar el informe de antenas: {e}")
        print("Detalles del error:", repr(e))
