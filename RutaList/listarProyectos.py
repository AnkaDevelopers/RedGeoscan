from pathlib import PureWindowsPath
from datetime import datetime
import pandas as pd
import os
import time
from openpyxl import load_workbook  # <-- agregado para ajustar anchos

# Importar módulos de monitor
from monitor.log.log import agregar_log

# --- Helpers de limpieza ---
def limpiar_ruta(ruta):
    if not ruta:
        return ruta
    ruta = str(ruta).replace('/', '\\')
    return str(PureWindowsPath(ruta))

def limpiar_fecha(fecha):
    if not fecha:
        return fecha
    try:
        f = str(fecha).replace('Z', '+00:00')
        dt = datetime.fromisoformat(f)
        return dt.date().isoformat()  # YYYY-MM-DD
    except Exception:
        s = str(fecha)[:10]
        try:
            return datetime.strptime(s, "%Y-%m-%d").date().isoformat()
        except Exception:
            return None

# --- Función principal ---
def listarProyectos(Response, rutaExcel):
    try:
        agregar_log("Iniciando listarProyectos")

        # 1) Asegurar lista
        if not isinstance(Response, list):
            Response = [Response] if Response else []
        agregar_log(f"Proyectos recibidos: {len(Response)}")

        # 1.1) Dedup por ID_PROYECTO
        vistos = set()
        normalizados = []
        for item in reversed(Response):
            if isinstance(item, dict):
                _id = item.get("ID_PROYECTO")
                if _id not in vistos:
                    vistos.add(_id)
                    normalizados.append(item)
        normalizados.reverse()
        agregar_log(f"Proyectos tras deduplicar: {len(normalizados)}")

        # 1.2) Limpiar ruta y fecha
        proyectos_limpios = []
        for it in normalizados:
            nuevo = it.copy()
            if "RUTA_PROYECTO" in nuevo:
                nuevo["RUTA_PROYECTO"] = limpiar_ruta(nuevo["RUTA_PROYECTO"])
            if "FECHA_CREACION" in nuevo:
                nuevo["FECHA_CREACION"] = limpiar_fecha(nuevo["FECHA_CREACION"])
            proyectos_limpios.append(nuevo)
        agregar_log("Datos limpiados")

        df_new = pd.DataFrame(proyectos_limpios)

        # 2) Columnas preferidas
        columnas_pref = [
            "ID_PROYECTO", "NOMBRE_PROYECTO", "FECHA_CREACION",
            "RUTA_PROYECTO", "RADIO_BUSQUEDA", "ESTADO_RED",
            "ESTADO_GEO", "NOMBRE_TOPOGRAFO", "NOMBRE_EMPRESA",
        ]

        # 3) Rutas de archivo (rutaExcel es la RUTA COMPLETA del .xlsx)
        ruta_final = rutaExcel
        ruta_dir = os.path.dirname(ruta_final)
        ruta_tmp = os.path.join(ruta_dir, "~" + os.path.basename(ruta_final))

        # 4) Crear archivo si no existe
        if not os.path.exists(ruta_final):
            agregar_log("Excel no existe, creando nuevo")
            cols_presentes = [c for c in columnas_pref if c in df_new.columns]
            if cols_presentes:
                df_export = df_new[cols_presentes].copy()
                extras = [c for c in df_new.columns if c not in cols_presentes]
                if extras:
                    df_export = pd.concat([df_export, df_new[extras]], axis=1)
            else:
                df_export = df_new
            df_export.to_excel(ruta_tmp, index=False)
            os.replace(ruta_tmp, ruta_final)
            # --- Ajuste de anchos ---
            try:
                wb = load_workbook(ruta_final)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value is not None:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                wb.save(ruta_final)
                agregar_log("Ajuste de anchos aplicado")
            except Exception as e:
                agregar_log(f"No se pudo ajustar anchos: {str(e)}")
            # -----------------------
            agregar_log("Excel creado desde cero")
            return True

        # 5) Actualizar si existe
        agregar_log("Excel existente, actualizando")
        df_old = pd.read_excel(ruta_final, engine="openpyxl")

        # 5.1) Tipos consistentes de ID
        if "ID_PROYECTO" in df_old.columns:
            df_old["ID_PROYECTO"] = pd.to_numeric(df_old["ID_PROYECTO"], errors="coerce")
        if "ID_PROYECTO" in df_new.columns:
            df_new["ID_PROYECTO"] = pd.to_numeric(df_new["ID_PROYECTO"], errors="coerce")

        # 5.2) Columnas protegidas
        cols_protegidas = {"ID_PROYECTO", "NOMBRE_TOPOGRAFO", "FECHA_CREACION", "RUTA_PROYECTO"}

        # 5.3) Unificar columnas
        todas = list(dict.fromkeys([*df_old.columns, *df_new.columns]))
        for c in todas:
            if c not in df_old.columns:
                df_old[c] = pd.NA
            if c not in df_new.columns:
                df_new[c] = pd.NA

        # 5.4) Verificar columnas de ID
        if "ID_PROYECTO" not in df_old.columns or "ID_PROYECTO" not in df_new.columns:
            agregar_log("No se encontró columna ID_PROYECTO, creando desde cero")
            df_new[todas].to_excel(ruta_tmp, index=False)
            os.replace(ruta_tmp, ruta_final)
            # --- Ajuste de anchos ---
            try:
                wb = load_workbook(ruta_final)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value is not None:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                wb.save(ruta_final)
                agregar_log("Ajuste de anchos aplicado")
            except Exception as e:
                agregar_log(f"No se pudo ajustar anchos: {str(e)}")
            # -----------------------
            return True

        df_old = df_old.set_index("ID_PROYECTO", drop=False)
        df_new = df_new.set_index("ID_PROYECTO", drop=False)

        ids_comunes = df_old.index.intersection(df_new.index)
        ids_nuevos  = df_new.index.difference(df_old.index)
        agregar_log(f"IDs comunes: {len(ids_comunes)}, IDs nuevos: {len(ids_nuevos)}")

        # 5.6) Columnas actualizables
        cols_actualizables = [c for c in todas if c not in cols_protegidas]

        # 5.7) Actualizar existentes
        if len(ids_comunes) > 0 and cols_actualizables:
            df_old.loc[ids_comunes, cols_actualizables] = df_new.loc[ids_comunes, cols_actualizables].values
            agregar_log("Filas existentes actualizadas")

        # 5.8) Agregar nuevos
        if len(ids_nuevos) > 0:
            df_old = pd.concat([df_old, df_new.loc[ids_nuevos, todas]], axis=0)
            agregar_log("Filas nuevas agregadas")

        # 5.9) Reordenar columnas
        cols_presentes = [c for c in columnas_pref if c in df_old.columns]
        df_export = df_old
        if cols_presentes:
            extras = [c for c in df_old.columns if c not in cols_presentes]
            df_export = pd.concat([df_old[cols_presentes], df_old[extras]], axis=1)

        # 6) Guardar
        df_export.to_excel(ruta_tmp, index=False)
        os.replace(ruta_tmp, ruta_final)
        # --- Ajuste de anchos ---
        try:
            wb = load_workbook(ruta_final)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(ruta_final)
            agregar_log("Ajuste de anchos aplicado")
        except Exception as e:
            agregar_log(f"No se pudo ajustar anchos: {str(e)}")
        # -----------------------
        agregar_log("Excel actualizado correctamente")

        return True

    except Exception as e:
        agregar_log(f"Error en listarProyectos: {str(e)}")
        return None
