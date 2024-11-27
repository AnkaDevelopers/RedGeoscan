from calculos import decimales_a_gms
from openpyxl.styles import PatternFill
import openpyxl
import os

ruta_principal = None

def insertar_datos_antenas(tabla_antenas, antenas_con_rinex, ruta):
    global ruta_principal
    ruta_principal = ruta

    print(f"Ruta principal: {ruta_principal}")

    # Limpiar la tabla en la interfaz
    for item in tabla_antenas.get_children():
        tabla_antenas.delete(item)

    # Filtrar antenas que tienen has_rinex=True
    antenas_filtradas = antenas_con_rinex[antenas_con_rinex['has_rinex'] == True]

    # Crear una lista para exportar al Excel (toda la lista completa)
    datos_para_exportar = []

    # Insertar las antenas filtradas en la tabla y preparar datos para el Excel
    for _, antena in antenas_filtradas.iterrows():
        nombre = antena['NAME']
        lat_g, lat_m, lat_s = decimales_a_gms(antena['Latitud (Decimal)'])
        lon_g, lon_m, lon_s = decimales_a_gms(antena['Longitud (Decimal)'])
        distancia = f"{round(antena['Distancia'], 2)} km"
        administrador = antena['ADMINISTRADOR']

        # Insertar en la tabla del programa
        tabla_antenas.insert("", "end", values=(
            nombre,
            f"{lat_g}° {lat_m}' {lat_s:.2f}\"",
            f"{lon_g}° {lon_m}' {lon_s:.2f}\"",
            distancia,
            administrador,
        ))

    # Preparar datos para exportar al Excel (toda la lista, incluyendo True y False en has_rinex)
    for _, antena in antenas_con_rinex.iterrows():
        nombre = antena['NAME']
        lat_g, lat_m, lat_s = decimales_a_gms(antena['Latitud (Decimal)'])
        lon_g, lon_m, lon_s = decimales_a_gms(antena['Longitud (Decimal)'])
        distancia = f"{round(antena['Distancia'], 2)} km"
        administrador = antena['ADMINISTRADOR']
        rinex_encontrado = "Sí" if antena['has_rinex'] else "No"

        # Agregar datos al Excel
        datos_para_exportar.append([
            nombre,
            f"{lat_g}° {lat_m}' {lat_s:.2f}\"",
            f"{lon_g}° {lon_m}' {lon_s:.2f}\"",
            distancia,
            administrador,
            rinex_encontrado
        ])

    # Exportar los datos a un archivo Excel
    exportar_datos_a_excel(datos_para_exportar, ruta_principal)

def exportar_datos_a_excel(datos, ruta_base):
    try:
        # Crear la ruta de la carpeta 'lista-control-antenas'
        ruta_carpeta = os.path.join(ruta_base, "1-lista-control-antenas")

        # Crear la carpeta si no existe
        if not os.path.exists(ruta_carpeta):
            os.makedirs(ruta_carpeta)
            print(f"Carpeta creada: {ruta_carpeta}")
        else:
            print(f"Carpeta ya existente: {ruta_carpeta}")

        # Definir la ruta completa del archivo Excel
        archivo_excel = os.path.join(ruta_carpeta, "lista_de_antenas.xlsx")

        # Crear un libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Antenas Más Cercanas"

        # Escribir encabezados
        encabezados = ["Nombre", "Latitud", "Longitud", "Distancia", "Administrador", "RINEX Encontrado"]
        ws.append(encabezados)

        # Crear estilos para las filas
        fill_verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
        fill_rojo = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Rojo claro

        # Escribir los datos y aplicar estilos
        for fila in datos:
            ws.append(fila)

            # Obtener el índice de la fila actual
            fila_actual = ws.max_row
            valor_rinex = fila[-1]  # Última columna es "RINEX Encontrado"

            # Seleccionar estilo según el valor de "RINEX Encontrado"
            if valor_rinex == "Sí":
                fill = fill_verde
            else:
                fill = fill_rojo

            # Aplicar el estilo a toda la fila
            for col in range(1, len(fila) + 1):  # Desde la columna 1 hasta el número de columnas en la fila
                ws.cell(row=fila_actual, column=col).fill = fill

        # Ajustar el ancho de las columnas a 130 píxeles
        for col in ws.columns:
            column_letter = col[0].column_letter  # Obtener la letra de la columna
            ws.column_dimensions[column_letter].width = 20  # 130px ≈ 20 caracteres de ancho

        # Guardar el archivo en la ruta especificada
        wb.save(archivo_excel)
        print(f"Datos exportados exitosamente a: {archivo_excel}")

    except Exception as e:
        print(f"Error al exportar los datos a Excel: {e}")
        print("Detalles del error:", repr(e))
